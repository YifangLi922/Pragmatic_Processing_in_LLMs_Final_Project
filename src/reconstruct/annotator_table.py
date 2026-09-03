"""Module 1: load one annotator's raw answer table (plan section 4.2).

Real file: "SFP母语者标注N ....xlsx", sheet "母语者填写". Columns:

    题号                                          shuffled_index (join key)
    情景 / 句子 / 问题                              content shown to the annotator
    A / B / C / D                                 option texts (should match the
                                                   master table's for the same item)
    母语者认为的正确答案                            answer_letter
    自然度（1–5）                                   naturalness
    是否犹豫（如无可留空）                          hesitation (any non-empty value = 1)
    如有犹豫：在哪几个选项之间，以及为什么           hesitation_reason
    是否觉得没有合适答案（如无可留空）              no_valid_option (any non-empty value = 1)
    若有，请简单写明你认为的正确答案                no_valid_option_detail
"""

from openpyxl import load_workbook


def _flag(value) -> int:
    return 1 if value not in (None, "") else 0


def _clean_letter(value) -> str | None:
    """Normalize stray whitespace/case in the answer-letter cell (e.g. "D ")
    -- a real typo we hit in the actual data -- without silently accepting
    genuinely invalid values.
    """
    if value is None:
        return None
    letter = str(value).strip().upper()
    return letter if letter in ("A", "B", "C", "D") else (letter or None)


def parse_annotator_rows(rows: list[tuple]) -> dict[str, dict]:
    """`rows` = raw value tuples from a "母语者填写" sheet, header row excluded.
    Returns {shuffled_index: record}.
    """
    records = {}
    for row in rows:
        shuffled_index = row[0]
        if not shuffled_index:
            continue
        records[shuffled_index] = {
            "shuffled_index": shuffled_index,
            "context": row[1],
            "sentence": row[2],
            "question": row[3],
            "options": {"A": row[4], "B": row[5], "C": row[6], "D": row[7]},
            "answer_letter": _clean_letter(row[8]),
            "naturalness": row[9] if isinstance(row[9], (int, float)) else None,
            "hesitation": _flag(row[10]),
            "hesitation_reason": row[11] or "",
            "no_valid_option": _flag(row[12]),
            "no_valid_option_detail": row[13] or "",
        }
    return records


def load_annotator_table(path: str, sheet: str = "母语者填写") -> dict[str, dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    return parse_annotator_rows(rows)
