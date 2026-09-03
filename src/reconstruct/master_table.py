"""Module 1: load the master answer-key table (plan section 4.1).

Real file: "SFP标注完整版.xlsx", sheet "研究者答案键". Real column names differ
from the plan's placeholder English schema but the structure is the same one
row per item, keyed by shuffled_index ("题号"):

    题号            shuffled_index, e.g. "Q001" (join key against annotator tables)
    Family          "Family 18" -> normalized to family_id "F18"
    来源            provenance note (原20题 / 新增16题（修改版）) -- kept as metadata
    源文档题号       item's number within its own source batch -- kept as metadata
    条件            particle_condition, Chinese ("裸句"/"+吧"/"+吗") -> bare/ba/ma
    句子            sentence
    设计预期答案     gold_letter_designed
    设计预期含义     free-text description of the designed-gold option's role
                    (not used directly -- option_semantics is derived from
                    the A-D option text instead, see semantics.py)
    A/B/C/D         option texts

Columns beyond D (in this file: a blank spacer column, then a side summary
panel reusing the same sheet) are ignored.
"""

import re

from openpyxl import load_workbook

from .semantics import LETTERS, derive_option_semantics

CONDITION_MAP = {"裸句": "bare", "+吧": "ba", "+吗": "ma"}

_HEADER = (
    "题号",
    "Family",
    "来源",
    "源文档题号",
    "条件",
    "句子",
    "设计预期答案",
    "设计预期含义",
    "A",
    "B",
    "C",
    "D",
)


def _family_id(raw: str) -> str:
    m = re.search(r"\d+", raw or "")
    if not m:
        raise ValueError(f"can't find a family number in {raw!r}")
    return f"F{int(m.group()):02d}"


def parse_master_rows(rows: list[tuple]) -> dict[str, dict]:
    """`rows` = raw value tuples from the answer-key sheet, header row excluded.
    Returns {shuffled_index: item_meta}. Rows with an empty 题号 (e.g. the
    side summary panel's rows, which reuse later columns of this same sheet)
    are skipped.
    """
    items = {}
    for row in rows:
        shuffled_index = row[0]
        if not shuffled_index:
            continue
        condition_raw = row[4]
        option_texts = {letter: row[8 + i] for i, letter in enumerate(LETTERS)}
        option_semantics, warnings = derive_option_semantics(option_texts)

        items[shuffled_index] = {
            "shuffled_index": shuffled_index,
            "family_id": _family_id(row[1]),
            "source": row[2],
            "source_item_no": row[3],
            "particle_condition": CONDITION_MAP.get(condition_raw, condition_raw),
            "sentence": row[5],
            "gold_letter_designed": row[6],
            "options": option_texts,
            "option_semantics": option_semantics,
            "option_semantics_warnings": warnings,
        }
    return items


def load_master_table(path: str, sheet: str = "研究者答案键") -> dict[str, dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    return parse_master_rows(rows)
